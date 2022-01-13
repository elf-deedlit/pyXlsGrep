#!/usr/bin/env python3
# vim: set ts=4 sw=4 et smartindent ignorecase fileencoding=utf8:
import argparse
import re
import os
# https://openpyxl.readthedocs.io/en/stable/
import openpyxl
import shutil
import unicodedata

from openpyxl.utils.exceptions import InvalidFileException

# 条件付き書式が設定されていると出るメッセージを抑制する
# UserWarning: Conditional Formatting extension is not supported and will be removed
import warnings
warnings.filterwarnings('ignore', 'Conditional Formatting.*', category=UserWarning)

BASEPATH = os.path.dirname(os.path.abspath(__file__))
FNMATCH = re.compile('(?s:.*\.xls[xm]?)$')
COLUMNS = shutil.get_terminal_size().columns - 1

def in_value(fs: str, value: str) -> bool:
    if not isinstance(value, str):
        value = repr(value)
    return fs.lower() in value.lower()

def search_xlsx(filename: str, fs: str) -> bool:
    try:
        # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.reader.excel.html?highlight=load_workbook
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
    except InvalidFileException:
        print(f'{filename}の形式はサポートしていません')
        return
    except PermissionError as err:
        print(f'{filename}が開けません。: {err.strerror}({err.errno})')
        return
    nocrlf = False
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and in_value(fs, cell.value):
                    if nocrlf is False:
                        print('')   # 改行のみさせたい
                        print(f'find: {filename}')
                        nocrlf = True
                    print(f'[{sheetname}]({cell.coordinate})={cell.value}')
    return nocrlf

def column_cut_msg(msg: str, width: int = COLUMNS) -> str:
    # ターミナルの横幅まで文字列を切り詰める
    # textwrapperは日本語に対応していない
    msg = unicodedata.normalize('NFC', msg)
    count = 0
    rslt = []
    for c in msg[::-1]:
        if unicodedata.east_asian_width(c) in 'FWA':
            count += 2
        else:
            count += 1
        if count >= width:
            rslt.reverse()
            t = ''.join(rslt)
            return '…' + t
        rslt.append(c)
    return msg

def find_xls(path: str, fs: str) -> None:
    for root, _, files in os.walk(path):
        for f in files:
            if re.match(FNMATCH, f):
                fullpath = os.path.join(root, f)
                msg = column_cut_msg(f)
                print(f'{msg}', end='\r', flush=True)
                crlf = search_xlsx(fullpath, fs)
                if crlf is False:
                    print(' '*COLUMNS, end='\r', flush=True)

def option_parse() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='エクセルファイルをgrepする')
    parser.add_argument('--basepath', type=str, default=BASEPATH, help='検索パス')
    parser.add_argument('findstr', type=str, help='検索文字列')
    return parser.parse_args()

def main() -> None:
    args = option_parse()
    find_xls(args.basepath, args.findstr)

if __name__ == '__main__':
    main()